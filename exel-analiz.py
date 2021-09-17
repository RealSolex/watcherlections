import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from browser import driver
driver = driver()
driver.download_raspisanie()
driver.close_driver(3)
wb = openpyxl.load_workbook('ИИТ_2 курс_21-22_осень.xlsx')
sheet = wb.active
sheet.cell()